"""
scripts/generate_test_data.py
Generates test_data/test_data.xlsx with all test data sheets.
Run once before test execution:
    python scripts/generate_test_data.py
"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


OUTPUT_DIR  = os.path.join(os.path.dirname(os.path.dirname(__file__)), "test_data")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "test_data.xlsx")

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col)
        c.value = h
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border    = thin_border()
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = max(len(h) + 4, 18)

    for row_idx, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            c = ws.cell(row_idx, col)
            c.value     = val
            c.border    = thin_border()
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"


def build_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── AUTH_REGISTER ─────────────────────────────────────────────────────────
    write_sheet(wb, "AUTH_REGISTER",
        headers=["tc_id", "email", "first_name", "last_name", "password",
                 "agree_terms", "expected_result", "note"],
        rows=[
            ["TC_REG_24", "testuser_sel_01@test.com", "Selenium", "User",
             "Abc@1234", "true", "PASS - Redirect to Dashboard", "Happy path"],
            ["TC_REG_25", "existing@test.com", "Test", "User",
             "Abc@1234", "true", "FAIL - Email already exists", "Duplicate email"],
            ["TC_REG_09", "", "Test", "User",
             "Abc@1234", "true", "FAIL - First name required", "Empty first name"],
            ["TC_REG_10", "valid@test.com", "", "",
             "Abc@1234", "true", "FAIL - Name required", "Empty name"],
            ["TC_REG_11", "valid@test.com", "Test", "User",
             "", "true", "FAIL - Password required", "Empty password"],
            ["TC_REG_12", "", "", "",
             "", "false", "FAIL - All fields required", "All empty"],
            ["TC_REG_13", "usertest.com", "Test", "User",
             "Abc@1234", "true", "FAIL - Invalid email", "Missing @"],
            ["TC_REG_14", "user@", "Test", "User",
             "Abc@1234", "true", "FAIL - Invalid email", "Missing domain"],
            ["TC_REG_15", "user @test.com", "Test", "User",
             "Abc@1234", "true", "FAIL - Invalid email", "Space in email"],
            ["TC_REG_16", "valid@test.com", "Test", "User",
             "Abc@12", "true", "FAIL - Password too short", "< 8 chars"],
            ["TC_REG_17", "valid@test.com", "Test", "User",
             "abc@1234", "true", "FAIL - Missing uppercase", "No uppercase"],
            ["TC_REG_18", "valid@test.com", "Test", "User",
             "ABC@1234", "true", "FAIL - Missing lowercase", "No lowercase"],
            ["TC_REG_19", "valid@test.com", "Test", "User",
             "Abcdefg@", "true", "FAIL - Missing digit", "No digit"],
            ["TC_REG_20", "valid@test.com", "Test", "User",
             "Abcde123", "true", "FAIL - Missing special char", "No special char"],
            ["TC_REG_26", "chartest_sel@test.com", "Char", "Test",
             "Hello@World9", "true", "PASS", "All 4 char types"],
            ["TC_REG_27", "test.user_01@test.com", "Dot", "User",
             "Abc@1234", "true", "PASS", "Email with dots/underscores"],
        ]
    )

    # ── AUTH_LOGIN ────────────────────────────────────────────────────────────
    write_sheet(wb, "AUTH_LOGIN",
        headers=["tc_id", "email", "password", "expected_result", "note"],
        rows=[
            ["TC_LOGIN_10", "active@test.com", "Abc@1234",
             "PASS - Redirect to Dashboard", "Valid credentials"],
            ["TC_LOGIN_11", "active@test.com", "WrongPass@99",
             "FAIL - Wrong password", "Incorrect password"],
            ["TC_LOGIN_12", "notexist@test.com", "Abc@1234",
             "FAIL - Email not found", "Unregistered email"],
            ["TC_LOGIN_13", "locked@test.com", "Abc@1234",
             "FAIL - Account locked", "Locked account"],
            ["TC_LOGIN_07", "", "Abc@1234",
             "FAIL - Email required", "Empty email"],
            ["TC_LOGIN_08", "active@test.com", "",
             "FAIL - Password required", "Empty password"],
            ["TC_LOGIN_09", "", "",
             "FAIL - Both required", "Both empty"],
            ["TC_LOGIN_16", "ACTIVE@TEST.COM", "Abc@1234",
             "PASS - Case insensitive", "Uppercase email"],
            ["TC_LOGIN_17", "active@test.com", "abc@1234",
             "FAIL - Case sensitive password", "Lowercase first char"],
        ]
    )

    # ── TASK ──────────────────────────────────────────────────────────────────
    write_sheet(wb, "TASK",
        headers=["tc_id", "title", "description", "status", "priority",
                 "expected_result", "note"],
        rows=[
            ["TC_TASK_01", "SEL_AUTO_Task_001", "Automation test task",
             "todo", "medium", "PASS - Task created", "Happy path"],
            ["TC_TASK_02", "", "No title task",
             "todo", "low", "FAIL - Title required", "Empty title"],
            ["TC_TASK_03", "T"*201, "Too long title",
             "todo", "low", "FAIL - Title too long", "> 200 chars"],
            ["TC_TASK_04", "SEL_AUTO_Task_002", "Search test",
             "todo", "high", "PASS - Found in search", "Search test"],
        ]
    )

    # ── GROUP ─────────────────────────────────────────────────────────────────
    write_sheet(wb, "GROUP",
        headers=["tc_id", "group_name", "description",
                 "expected_result", "note"],
        rows=[
            ["TC_GROUP_01", "SEL_AUTO_Group_001", "Selenium test group",
             "PASS - Group created", "Happy path"],
            ["TC_GROUP_02", "", "No name group",
             "FAIL - Name required", "Empty name"],
            ["TC_GROUP_03", "G"*257, "Too long name",
             "FAIL - Name too long", "> 256 chars"],
        ]
    )

    # ── USER_PROFILE ──────────────────────────────────────────────────────────
    write_sheet(wb, "USER_PROFILE",
        headers=["tc_id", "field", "value", "expected_result", "note"],
        rows=[
            ["TC_USER_01", "name", "Updated Selenium User",
             "PASS - Name updated", "Update name"],
            ["TC_USER_02", "name", "A"*101,
             "FAIL - Name too long", "> 100 chars"],
            ["TC_USER_03", "password", "NewAbc@9999",
             "PASS - Password changed", "Change password"],
        ]
    )

    # ── ADMIN ─────────────────────────────────────────────────────────────────
    write_sheet(wb, "ADMIN",
        headers=["tc_id", "action", "target_email",
                 "expected_result", "note"],
        rows=[
            ["TC_ADMIN_01", "access_admin", "active@test.com",
             "FAIL - Access denied", "Regular user blocked"],
            ["TC_ADMIN_02", "access_admin", "admin@test.com",
             "PASS - Admin panel accessible", "Admin user allowed"],
            ["TC_ADMIN_03", "view_users", "admin@test.com",
             "PASS - User list visible", "View all users"],
            ["TC_ADMIN_04", "lock_unlock", "active@test.com",
             "PASS - Lock/Unlock works", "Toggle lock state"],
        ]
    )

    # ── NFR ───────────────────────────────────────────────────────────────────
    write_sheet(wb, "NFR",
        headers=["tc_id", "nfr_id", "category", "description",
                 "threshold", "note"],
        rows=[
            ["TC_NFR_PERF_01", "NFR-PERF-1", "Performance",
             "Login response time", "< 3 seconds", ""],
            ["TC_NFR_PERF_02", "NFR-PERF-1", "Performance",
             "API health check", "< 3 seconds", ""],
            ["TC_NFR_PERF_03", "NFR-PERF-1", "Performance",
             "Page load time", "< 3 seconds", ""],
            ["TC_NFR_SEC_01", "NFR-SEC-2", "Security",
             "Unauthenticated redirect", "Redirect to login", ""],
            ["TC_NFR_SEC_02", "NFR-SEC-1", "Security",
             "Password not in DOM", "No plaintext in source", ""],
            ["TC_NFR_USAB_01", "NFR-USAB-1", "Usability",
             "Desktop responsive 1920x1080", "No layout break", ""],
            ["TC_NFR_USAB_02", "NFR-USAB-1", "Usability",
             "Tablet responsive 768x1024", "No layout break", ""],
            ["TC_NFR_USAB_03", "NFR-USAB-1", "Usability",
             "Mobile responsive 375x667", "No layout break", ""],
            ["TC_NFR_USAB_04", "NFR-USAB-4", "Usability",
             "User-friendly error messages", "No technical codes", ""],
        ]
    )

    return wb


if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT_FILE)
    print(f"✅ Test data generated → {OUTPUT_FILE}")
    print(f"   Sheets: {wb.sheetnames}")
