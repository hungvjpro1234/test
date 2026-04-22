"""
scripts/generate_report.py
Post-run report generator — reads pytest JSON output and writes a polished Excel + HTML summary.
Usage:
    pytest --json-report --json-report-file=reports/report.json ...
    python scripts/generate_report.py
"""
import os
import sys
import json
import datetime
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORT_DIR   = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")
JSON_FILE    = os.path.join(REPORT_DIR, "report.json")
OUTPUT_EXCEL = os.path.join(REPORT_DIR, f"final_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
OUTPUT_HTML  = os.path.join(REPORT_DIR, f"final_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.html")

PASS_COLOR   = "C6EFCE"
FAIL_COLOR   = "FFC7CE"
SKIP_COLOR   = "FFEB9C"
HEADER_COLOR = "1F3864"


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def status_fill(outcome):
    m = {"passed": PASS_COLOR, "failed": FAIL_COLOR,
         "skipped": SKIP_COLOR, "error": FAIL_COLOR}
    return PatternFill("solid", fgColor=m.get(outcome, "FFFFFF"))


def load_json_report():
    if not os.path.exists(JSON_FILE):
        print(f"⚠ JSON report not found at: {JSON_FILE}")
        print("  Run: pytest --json-report --json-report-file=reports/report.json")
        return None
    with open(JSON_FILE) as f:
        return json.load(f)


def build_excel_report(data):
    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = "📊 Summary"

    tests = data.get("tests", [])
    total   = len(tests)
    passed  = sum(1 for t in tests if t["outcome"] == "passed")
    failed  = sum(1 for t in tests if t["outcome"] == "failed")
    skipped = sum(1 for t in tests if t["outcome"] == "skipped")
    duration = data.get("duration", 0)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    summary_ws.cell(1, 1).value = "Test Execution Report"
    summary_ws.cell(1, 1).font  = Font(bold=True, size=14)
    summary_ws.merge_cells("A1:B1")

    def write_kv(row, k, v, highlight=None):
        c1 = summary_ws.cell(row, 1)
        c2 = summary_ws.cell(row, 2)
        c1.value = k
        c2.value = v
        c1.font = Font(bold=True)
        if highlight:
            c2.fill = PatternFill("solid", fgColor=highlight)
        for c in [c1, c2]:
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center")

    write_kv(2, "Run Date",   datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
    write_kv(3, "Duration",   f"{duration:.1f}s")
    write_kv(4, "Total Tests", total)
    write_kv(5, "PASSED",  passed,  PASS_COLOR)
    write_kv(6, "FAILED",  failed,  FAIL_COLOR if failed > 0 else None)
    write_kv(7, "SKIPPED", skipped, SKIP_COLOR if skipped > 0 else None)
    write_kv(8, "Pass Rate",
              f"{(passed/total*100):.1f}%" if total else "N/A",
              PASS_COLOR if total and passed/total >= 0.9 else FAIL_COLOR)
    summary_ws.column_dimensions["A"].width = 18
    summary_ws.column_dimensions["B"].width = 22

    # ── Details sheet ─────────────────────────────────────────────────────────
    det = wb.create_sheet("📋 Test Results")
    headers = ["Test ID", "Module", "Test Name", "Outcome",
               "Duration (s)", "Error Message"]
    for col, h in enumerate(headers, 1):
        c = det.cell(1, col)
        c.value     = h
        c.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
        c.font      = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        c.border    = thin_border()

    col_widths = [16, 22, 45, 10, 12, 50]
    for i, w in enumerate(col_widths, 1):
        det.column_dimensions[get_column_letter(i)].width = w

    for row_idx, t in enumerate(tests, 2):
        nodeid    = t.get("nodeid", "")
        outcome   = t.get("outcome", "")
        dur       = t.get("duration", 0)
        # Parse module from nodeid
        parts     = nodeid.split("::")
        module    = parts[1] if len(parts) > 1 else ""
        test_name = parts[-1] if parts else nodeid
        tc_id     = test_name.split("_test_")[-1] if "_test_" in test_name else test_name[:20]
        error_msg = ""
        if t.get("call") and t["call"].get("longrepr"):
            error_msg = str(t["call"]["longrepr"])[:200]

        cells_data = [tc_id, module, test_name, outcome.upper(), f"{dur:.2f}", error_msg]
        for col, val in enumerate(cells_data, 1):
            c = det.cell(row_idx, col)
            c.value     = val
            c.border    = thin_border()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 4:
                c.fill      = status_fill(outcome)
                c.font      = Font(bold=True)
                c.alignment = Alignment(horizontal="center")

    det.freeze_panes = "A2"
    det.auto_filter.ref = f"A1:F{len(tests)+1}"

    wb.save(OUTPUT_EXCEL)
    print(f"\n📊 Excel report → {OUTPUT_EXCEL}")
    return wb


def build_html_report(data):
    tests    = data.get("tests", [])
    total    = len(tests)
    passed   = sum(1 for t in tests if t["outcome"] == "passed")
    failed   = sum(1 for t in tests if t["outcome"] == "failed")
    skipped  = sum(1 for t in tests if t["outcome"] == "skipped")
    duration = data.get("duration", 0)
    pass_rate = f"{(passed/total*100):.1f}%" if total else "N/A"

    rows_html = ""
    for t in tests:
        outcome   = t.get("outcome", "")
        nodeid    = t.get("nodeid", "")
        dur       = t.get("duration", 0)
        parts     = nodeid.split("::")
        module    = parts[1] if len(parts) > 1 else ""
        test_name = parts[-1] if parts else nodeid
        color_map = {"passed": "#C6EFCE", "failed": "#FFC7CE",
                     "skipped": "#FFEB9C"}
        bg = color_map.get(outcome, "#fff")
        rows_html += f"""
        <tr>
          <td style="font-family:monospace;font-size:12px">{test_name}</td>
          <td>{module}</td>
          <td style="background:{bg};font-weight:bold;text-align:center">
            {outcome.upper()}</td>
          <td style="text-align:right">{dur:.2f}s</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Selenium Test Report — Todo List App</title>
<style>
  body {{ font-family: 'Segoe UI', sans-serif; margin: 0; background: #f4f6f9; color: #333; }}
  .header {{ background: linear-gradient(135deg,#1a3762,#2d5fa6); color: #fff;
             padding: 32px 40px; }}
  .header h1 {{ margin: 0; font-size: 26px; }}
  .header p  {{ margin: 4px 0 0; opacity: .8; font-size: 14px; }}
  .summary   {{ display: flex; gap: 16px; padding: 24px 40px; flex-wrap: wrap; }}
  .card      {{ background: #fff; border-radius: 12px; padding: 20px 28px;
                box-shadow: 0 2px 8px rgba(0,0,0,.08); min-width: 120px; text-align: center; }}
  .card .num {{ font-size: 36px; font-weight: 700; }}
  .card .lbl {{ font-size: 13px; color: #888; text-transform: uppercase; letter-spacing: 1px; }}
  .pass  {{ color: #1e7e34; }}
  .fail  {{ color: #c82333; }}
  .skip  {{ color: #e0a800; }}
  .total {{ color: #1a3762; }}
  table  {{ width: calc(100% - 80px); margin: 0 40px 40px; border-collapse: collapse;
            background: #fff; border-radius: 10px; overflow: hidden;
            box-shadow: 0 2px 8px rgba(0,0,0,.08); }}
  th     {{ background: #1a3762; color: #fff; padding: 12px 16px;
            font-size: 13px; text-align: left; }}
  td     {{ padding: 10px 16px; border-bottom: 1px solid #f0f0f0; font-size: 13px; }}
  tr:last-child td {{ border-bottom: none; }}
</style>
</head>
<body>
<div class="header">
  <h1>🧪 Selenium System Test Report</h1>
  <p>Project: Todo List App &nbsp;|&nbsp; Run: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}
     &nbsp;|&nbsp; Duration: {duration:.1f}s</p>
</div>
<div class="summary">
  <div class="card"><div class="num total">{total}</div><div class="lbl">Total</div></div>
  <div class="card"><div class="num pass">{passed}</div><div class="lbl">Passed</div></div>
  <div class="card"><div class="num fail">{failed}</div><div class="lbl">Failed</div></div>
  <div class="card"><div class="num skip">{skipped}</div><div class="lbl">Skipped</div></div>
  <div class="card"><div class="num {'pass' if passed==total else 'fail'}">{pass_rate}</div>
    <div class="lbl">Pass Rate</div></div>
</div>
<table>
  <thead>
    <tr><th>Test Name</th><th>Module</th><th>Outcome</th><th>Duration</th></tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>
</body>
</html>"""

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"🌐 HTML  report → {OUTPUT_HTML}")


if __name__ == "__main__":
    os.makedirs(REPORT_DIR, exist_ok=True)
    data = load_json_report()
    if data:
        build_excel_report(data)
        build_html_report(data)
