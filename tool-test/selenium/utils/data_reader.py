"""
data_reader.py — Read test data from Excel (.xlsx)
"""
import openpyxl
from utils.config import Config


class DataReader:
    """
    Reads rows from an Excel sheet.
    The first row is treated as headers (case-insensitive).
    Usage:
        reader = DataReader()
        rows   = reader.get_sheet("AUTH_REGISTER")   # → list[dict]
    """

    def __init__(self, path: str = None):
        self.path = path or Config.TEST_DATA_FILE
        self._wb = None

    def _load(self):
        if self._wb is None:
            self._wb = openpyxl.load_workbook(self.path, data_only=True)

    def get_sheet(self, sheet_name: str) -> list[dict]:
        self._load()
        if sheet_name not in self._wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {self.path}. "
                             f"Available: {self._wb.sheetnames}")
        ws = self._wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if all(c is None for c in row):
                continue          # skip completely empty rows
            result.append({headers[i]: row[i] for i in range(len(headers))})
        return result

    def get_row(self, sheet_name: str, filter_key: str, filter_val):
        for row in self.get_sheet(sheet_name):
            if str(row.get(filter_key.lower(), "")).strip() == str(filter_val).strip():
                return row
        return None

    def list_sheets(self) -> list[str]:
        self._load()
        return self._wb.sheetnames
