"""
report_helper.py — Excel/CSV report generation for test results
"""
import os
import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


class ReportHelper:
    """Generates an Excel test report with per-sheet module results."""

    PASS_COLOR  = "C6EFCE"   # green
    FAIL_COLOR  = "FFC7CE"   # red
    SKIP_COLOR  = "FFEB9C"   # yellow
    HEADER_COLOR = "1F3864"  # dark blue
    SUBHEADER_COLOR = "2F5496"

    def __init__(self):
        self.results: list[dict] = []
        self.start_time = datetime.datetime.now()

    def record(self, tc_id: str, module: str, description: str,
               status: str, actual: str = "", note: str = ""):
        """
        status: 'PASS' | 'FAIL' | 'SKIP'
        """
        self.results.append({
            "tc_id":       tc_id,
            "module":      module,
            "description": description,
            "status":      status,
            "actual":      actual,
            "note":        note,
            "timestamp":   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    def save(self, path: str = None):
        if path is None:
            ts = self.start_time.strftime("%Y%m%d_%H%M%S")
            path = os.path.join(
                os.path.dirname(__file__), "..", "reports",
                f"test_report_{ts}.xlsx"
            )
        os.makedirs(os.path.dirname(path), exist_ok=True)

        wb = openpyxl.Workbook()
        self._write_summary_sheet(wb)
        self._write_detail_sheet(wb)
        self._write_per_module_sheets(wb)

        wb.save(path)
        print(f"\n📊 Report saved → {path}")
        return path

    # ── Internal writers ──────────────────────────────────────────────────────
    def _header_style(self):
        return {
            "fill":      PatternFill("solid", fgColor=self.HEADER_COLOR),
            "font":      Font(bold=True, color="FFFFFF", size=11),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border":    self._thin_border(),
        }

    def _thin_border(self):
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def _apply_header(self, cell, value):
        cell.value = value
        style = self._header_style()
        cell.fill      = style["fill"]
        cell.font      = style["font"]
        cell.alignment = style["alignment"]
        cell.border    = style["border"]

    def _status_fill(self, status: str):
        colors = {"PASS": self.PASS_COLOR, "FAIL": self.FAIL_COLOR, "SKIP": self.SKIP_COLOR}
        return PatternFill("solid", fgColor=colors.get(status, "FFFFFF"))

    def _write_summary_sheet(self, wb: openpyxl.Workbook):
        ws = wb.active
        ws.title = "📊 Summary"

        total = len(self.results)
        passed = sum(1 for r in self.results if r["status"] == "PASS")
        failed = sum(1 for r in self.results if r["status"] == "FAIL")
        skipped = sum(1 for r in self.results if r["status"] == "SKIP")
        duration = (datetime.datetime.now() - self.start_time).seconds

        headers = ["Metric", "Value"]
        for col, h in enumerate(headers, 1):
            self._apply_header(ws.cell(1, col), h)

        rows = [
            ("Run Date",          self.start_time.strftime("%Y-%m-%d %H:%M")),
            ("Duration (s)",      duration),
            ("Total Test Cases",  total),
            ("PASS",              passed),
            ("FAIL",              failed),
            ("SKIP",              skipped),
            ("Pass Rate",         f"{(passed/total*100):.1f}%" if total else "N/A"),
        ]
        for i, (k, v) in enumerate(rows, 2):
            ws.cell(i, 1).value = k
            ws.cell(i, 2).value = v
            ws.cell(i, 1).font = Font(bold=True)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25

    def _write_detail_sheet(self, wb: openpyxl.Workbook):
        ws = wb.create_sheet("📋 All Results")
        headers = ["TC ID", "Module", "Description", "Status", "Actual Result", "Note", "Timestamp"]
        for col, h in enumerate(headers, 1):
            self._apply_header(ws.cell(1, col), h)
            ws.column_dimensions[get_column_letter(col)].width = [14, 16, 40, 8, 35, 25, 20][col-1]

        for row_idx, r in enumerate(self.results, 2):
            ws.cell(row_idx, 1).value = r["tc_id"]
            ws.cell(row_idx, 2).value = r["module"]
            ws.cell(row_idx, 3).value = r["description"]
            status_cell = ws.cell(row_idx, 4)
            status_cell.value = r["status"]
            status_cell.fill  = self._status_fill(r["status"])
            status_cell.font  = Font(bold=True)
            status_cell.alignment = Alignment(horizontal="center")
            ws.cell(row_idx, 5).value = r["actual"]
            ws.cell(row_idx, 6).value = r["note"]
            ws.cell(row_idx, 7).value = r["timestamp"]
            for col in range(1, 8):
                ws.cell(row_idx, col).border = self._thin_border()
                ws.cell(row_idx, col).alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:G{len(self.results)+1}"

    def _write_per_module_sheets(self, wb: openpyxl.Workbook):
        modules = {}
        for r in self.results:
            modules.setdefault(r["module"], []).append(r)

        for module, rows in modules.items():
            safe_name = module[:28].replace("/", "-").replace("\\", "-")
            ws = wb.create_sheet(f"📁 {safe_name}")
            headers = ["TC ID", "Description", "Status", "Actual Result", "Note"]
            for col, h in enumerate(headers, 1):
                self._apply_header(ws.cell(1, col), h)

            for row_idx, r in enumerate(rows, 2):
                ws.cell(row_idx, 1).value = r["tc_id"]
                ws.cell(row_idx, 2).value = r["description"]
                sc = ws.cell(row_idx, 3)
                sc.value = r["status"]
                sc.fill  = self._status_fill(r["status"])
                sc.font  = Font(bold=True)
                sc.alignment = Alignment(horizontal="center")
                ws.cell(row_idx, 4).value = r["actual"]
                ws.cell(row_idx, 5).value = r["note"]

            ws.column_dimensions["A"].width = 16
            ws.column_dimensions["B"].width = 45
            ws.column_dimensions["C"].width = 8
            ws.column_dimensions["D"].width = 35
            ws.column_dimensions["E"].width = 25
            ws.freeze_panes = "A2"
